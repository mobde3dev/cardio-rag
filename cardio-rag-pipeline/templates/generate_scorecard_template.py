"""
CardioRAG — Day 2 Retrieval Scorecard Template Generator (.xlsx)
===============================================================
Generates a professional Excel workbook with live formulas, KPI summary cards,
and conditional formatting to automatically compute Precision@5, Hit Rate, MRR,
and flag weak questions in red.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def create_retrieval_scorecard(output_xlsx_path: str):
    wb = openpyxl.Workbook()
    
    # Setup Sheet 1: Scorecard
    ws = wb.active
    ws.title = "Retrieval_Scorecard"
    ws.views.sheetView[0].showGridLines = True

    # Color Palette
    navy_dark = "1B365D"
    navy_header = "24436F"
    kpi_bg = "F0F4F9"
    kpi_border = "B0C4DE"
    white = "FFFFFF"
    gray_light = "F9FAFC"
    border_gray = "D3D3D3"
    green_soft = "E2F0D9"
    green_text = "385723"

    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color=navy_dark)
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="595959")
    font_kpi_val = Font(name="Segoe UI", size=16, bold=True, color=navy_dark)
    font_tbl_header = Font(name="Segoe UI", size=10, bold=True, color=white)
    font_data = Font(name="Segoe UI", size=9.5, color="000000")
    font_bold = Font(name="Segoe UI", size=9.5, bold=True, color="000000")

    # Fills & Borders
    fill_header = PatternFill(start_color=navy_header, end_color=navy_header, fill_type="solid")
    fill_kpi = PatternFill(start_color=kpi_bg, end_color=kpi_bg, fill_type="solid")
    fill_alt = PatternFill(start_color=gray_light, end_color=gray_light, fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color=border_gray),
        right=Side(style="thin", color=border_gray),
        top=Side(style="thin", color=border_gray),
        bottom=Side(style="thin", color=border_gray)
    )
    kpi_box_border = Border(
        left=Side(style="medium", color=kpi_border),
        right=Side(style="medium", color=kpi_border),
        top=Side(style="medium", color=kpi_border),
        bottom=Side(style="medium", color=kpi_border)
    )

    # 1. Title Block
    ws.merge_cells("A1:P1")
    ws["A1"] = "CARDIORAG — DAY 2 RETRIEVAL SCORECARD & BENCHMARK"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:P2")
    ws["A2"] = "Instructions: Enter 1 (Relevant) or 0 (Irrelevant) in Columns F-J for retrieved chunks. Precision@5, Hit Rate, and MRR calculate automatically. Weak questions are flagged in RED."
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = Alignment(vertical="center")

    # 2. KPI Summary Cards (Rows 4 to 6)
    kpi_cards = [
        ("B4:D4", "B5:D6", "TEAM AVERAGE PRECISION@5", "=AVERAGE(K9:K24)", "0.0%"),
        ("F4:H4", "F5:H6", "OVERALL HIT RATE (FOUND? %)", '=COUNTIF(L9:L24, "YES")/(COUNTA(C9:C24)-COUNTIF(D9:D24, "*expect refusal*"))', "0.0%"),
        ("J4:L4", "J5:L6", "REFUSAL ACCURACY %", '=COUNTIF(L9:L24, "Refusal Pass")/COUNTIF(D9:D24, "*expect refusal*")', "0.0%"),
        ("N4:P4", "N5:P6", "MEAN RECIPROCAL RANK (MRR)", "=AVERAGE(N9:N24)", "0.000"),
    ]

    for lbl_range, val_range, label, formula, num_format in kpi_cards:
        ws.merge_cells(lbl_range)
        top_left_lbl = lbl_range.split(":")[0]
        ws[top_left_lbl] = label
        ws[top_left_lbl].font = font_kpi_label
        ws[top_left_lbl].alignment = Alignment(horizontal="center", vertical="center")
        ws[top_left_lbl].fill = fill_kpi

        ws.merge_cells(val_range)
        top_left_val = val_range.split(":")[0]
        ws[top_left_val] = formula
        ws[top_left_val].font = font_kpi_val
        ws[top_left_val].alignment = Alignment(horizontal="center", vertical="center")
        ws[top_left_val].number_format = num_format
        ws[top_left_val].fill = fill_kpi

        # Apply borders to KPI cells
        cols = [val_range.split(":")[0][0], val_range.split(":")[1][0]]
        for col_idx in range(openpyxl.utils.column_index_from_string(cols[0]), openpyxl.utils.column_index_from_string(cols[1]) + 1):
            for row_idx in range(4, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = kpi_box_border
                cell.fill = fill_kpi

    # 3. Main Data Table Headers (Row 8)
    headers = [
        ("A8", "ID", 7),
        ("B8", "Guideline", 12),
        ("C8", "Evaluation Question", 45),
        ("D8", "Expected Target Source", 26),
        ("E8", "Expected Chunk ID", 22),
        ("F8", "Chunk 1", 9),
        ("G8", "Chunk 2", 9),
        ("H8", "Chunk 3", 9),
        ("I8", "Chunk 4", 9),
        ("J8", "Chunk 5", 9),
        ("K8", "Precision@5", 14),
        ("L8", "Found? (Hit@5)", 15),
        ("M8", "First Hit Rank", 14),
        ("N8", "Reciprocal Rank", 16),
        ("O8", "Spot-Checked?", 15),
        ("P8", "Facilitator Notes", 25),
    ]

    for cell_ref, text, col_width in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = font_tbl_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_letter = cell_ref[0] if len(cell_ref) == 2 else cell_ref[:2]
        ws.column_dimensions[col_letter].width = col_width

    # 4. Populate 16 Clinical Evaluation Rows (Rows 9 to 24)
    evaluation_questions = [
        # WHO 2021
        ("WHO_01", "WHO_2021", "What BP threshold should trigger starting medication?", "Recommendation 1, Page 3", "WHO03_3.1_REC_001", [1, 1, 0, 0, 0]),
        ("WHO_02", "WHO_2021", "Are laboratory tests required before starting pharmacological treatment?", "Recommendation 2, Page 5", "WHO03_3.2_REC_001", [1, 1, 0, 0, 0]),
        ("WHO_03", "WHO_2021", "When should cardiovascular disease risk assessment be conducted?", "Recommendation 3, Page 6", "WHO03_3.3_REC_001", [1, 0, 1, 0, 0]),
        ("WHO_04", "WHO_2021", "Which drug classes are recommended as first-line agents for hypertension?", "Recommendation 4, Page 7", "WHO03_3.4_REC_001", [1, 1, 0, 0, 0]),
        ("WHO_05", "WHO_2021", "When is combination therapy recommended as initial treatment?", "Recommendation 5, Page 8", "WHO03_3.5_REC_001", [1, 0, 0, 0, 0]),
        ("WHO_06", "WHO_2021", "What is the target BP for a patient with known CVD?", "Section 3.6, Page 9", "WHO03_3.6_REC_002", [1, 1, 0, 0, 0]),
        ("WHO_07", "WHO_2021", "Can nurses or pharmacists prescribe treatment?", "Section 3.8, Page 10", "WHO03_3.8_REC_001", [1, 1, 0, 0, 0]),
        ("WHO_08", "WHO_2021", "What's the recommended breast cancer screening interval?", "Not covered — expect refusal", "OUT_OF_SCOPE", [0, 0, 0, 0, 0]),
        
        # NICE 2023
        ("NICE_01", "NICE_2023", "What risk assessment tool is recommended for calculating 10-year CVD risk?", "Recommendation 1.1.7, Page 5", "NICE3_1.1.7_REC", [1, 1, 0, 0, 0]),
        ("NICE_02", "NICE_2023", "At what 10-year CVD risk threshold should atorvastatin 20 mg be offered for primary prevention?", "Recommendation 1.6.7, Page 21", "NICE3_1.6.7_REC", [1, 0, 0, 0, 0]),
        ("NICE_03", "NICE_2023", "What statin and dose is recommended for secondary prevention of CVD?", "Recommendation 1.7.2, Page 26", "NICE3_1.7.2_REC", [1, 1, 0, 0, 0]),
        ("NICE_04", "NICE_2023", "What is the target lipid level for secondary prevention of CVD?", "Recommendation 1.7.1, Page 26", "NICE3_1.7.1_REC", [1, 0, 0, 0, 0]),
        ("NICE_05", "NICE_2023", "What statin treatment is recommended for adults with chronic kidney disease (CKD)?", "Recommendation 1.8.1, Page 35", "NICE3_1.8.1_REC", [1, 1, 0, 0, 0]),
        ("NICE_06", "NICE_2023", "Should aspirin be routinely offered for primary prevention of CVD?", "Recommendation 1.2.1, Page 13", "NICE3_1.2.1_REC", [1, 0, 0, 0, 0]),
        ("NICE_07", "NICE_2023", "What treatment is recommended if statins are contraindicated or not tolerated?", "Recommendation 1.10.1, Page 38", "NICE3_1.10.1_REC", [1, 1, 0, 0, 0]),
        ("NICE_08", "NICE_2023", "What is the recommended antibiotic regimen for acute appendicitis?", "Not covered — expect refusal", "OUT_OF_SCOPE", [0, 0, 0, 0, 0]),
    ]

    for idx, (qid, gline, question, src, cid, sample_chunks) in enumerate(evaluation_questions, start=9):
        r_fill = fill_alt if idx % 2 == 0 else PatternFill(fill_type=None)
        
        ws[f"A{idx}"] = qid
        ws[f"A{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f"B{idx}"] = gline
        ws[f"B{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f"C{idx}"] = question
        ws[f"C{idx}"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws[f"D{idx}"] = src
        ws[f"D{idx}"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws[f"E{idx}"] = cid
        ws[f"E{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
        # 0/1 Chunk Relevance
        for c_offset, c_val in enumerate(sample_chunks):
            col_char = chr(ord("F") + c_offset)
            ws[f"{col_char}{idx}"] = c_val
            ws[f"{col_char}{idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col_char}{idx}"].font = font_bold

        # Live Formulas
        # K: Precision@5
        ws[f"K{idx}"] = f'=IF(ISNUMBER(SEARCH("expect refusal", D{idx})), IF(SUM(F{idx}:J{idx})=0, 1, 0), SUM(F{idx}:J{idx})/5)'
        ws[f"K{idx}"].number_format = "0.0%"
        ws[f"K{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"K{idx}"].font = font_bold

        # L: Hit@5 (Found?)
        ws[f"L{idx}"] = f'=IF(ISNUMBER(SEARCH("expect refusal", D{idx})), IF(SUM(F{idx}:J{idx})=0, "Refusal Pass", "Refusal Fail"), IF(SUM(F{idx}:J{idx})>0, "YES", "NO"))'
        ws[f"L{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"L{idx}"].font = font_bold

        # M: First Hit Rank
        ws[f"M{idx}"] = f'=IF(ISNUMBER(SEARCH("expect refusal", D{idx})), "N/A", IF(F{idx}=1, 1, IF(G{idx}=1, 2, IF(H{idx}=1, 3, IF(I{idx}=1, 4, IF(J{idx}=1, 5, 0))))))'
        ws[f"M{idx}"].alignment = Alignment(horizontal="center", vertical="center")

        # N: Reciprocal Rank
        ws[f"N{idx}"] = f'=IF(ISNUMBER(SEARCH("expect refusal", D{idx})), IF(SUM(F{idx}:J{idx})=0, 1, 0), IF(M{idx}>0, 1/M{idx}, 0))'
        ws[f"N{idx}"].number_format = "0.000"
        ws[f"N{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"N{idx}"].font = font_bold

        # O: Spot-Check Status
        ws[f"O{idx}"] = "Pending"
        ws[f"O{idx}"].alignment = Alignment(horizontal="center", vertical="center")

        # P: Facilitator Notes
        ws[f"P{idx}"] = ""

        # Apply borders and fonts
        for col_idx in range(1, 17):
            c = ws.cell(row=idx, column=col_idx)
            c.border = thin_border
            if c.fill.fill_type is None and idx % 2 == 0:
                c.fill = r_fill
            if col_idx not in [6, 7, 8, 9, 10, 11, 12, 14]:
                c.font = font_data

    # 5. Conditional Formatting Rules
    # Red flag for Precision@5 < 0.20 (Weak Question)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name="Segoe UI", size=9.5, bold=True, color="9C0006")
    rule_low_p = CellIsRule(operator="lessThan", formula=["0.20"], stopIfTrue=True, fill=red_fill, font=red_font)
    ws.conditional_formatting.add("K9:K24", rule_low_p)

    # Green flag for Precision@5 >= 0.40 (Strong Retrieval)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(name="Segoe UI", size=9.5, bold=True, color="006100")
    rule_high_p = CellIsRule(operator="greaterThanOrEqual", formula=["0.40"], stopIfTrue=True, fill=green_fill, font=green_font)
    ws.conditional_formatting.add("K9:K24", rule_high_p)

    # Hit@5 highlights
    rule_hit_no = CellIsRule(operator="equal", formula=['"NO"'], stopIfTrue=True, fill=red_fill, font=red_font)
    rule_hit_yes = CellIsRule(operator="equal", formula=['"YES"'], stopIfTrue=True, fill=green_fill, font=green_font)
    rule_refusal_pass = CellIsRule(operator="equal", formula=['"Refusal Pass"'], stopIfTrue=True, fill=green_fill, font=green_font)
    rule_refusal_fail = CellIsRule(operator="equal", formula=['"Refusal Fail"'], stopIfTrue=True, fill=red_fill, font=red_font)
    
    ws.conditional_formatting.add("L9:L24", rule_hit_no)
    ws.conditional_formatting.add("L9:L24", rule_hit_yes)
    ws.conditional_formatting.add("L9:L24", rule_refusal_pass)
    ws.conditional_formatting.add("L9:L24", rule_refusal_fail)

    # Setup Sheet 2: Facilitator Spot-Check Log
    ws_log = wb.create_sheet(title="Facilitator_SpotCheck_Log")
    ws_log.views.sheetView[0].showGridLines = True

    ws_log.merge_cells("A1:G1")
    ws_log["A1"] = "DAY 2 FACILITATOR SPOT-CHECK AUDIT LOG"
    ws_log["A1"].font = font_title

    ws_log.merge_cells("A2:G2")
    ws_log["A2"] = "Mandatory in-person audit: Facilitator randomly selects at least 1 real query per team to verify live retrieval against reported scorecard."
    ws_log["A2"].font = font_subtitle

    log_headers = [
        ("A4", "Team / Table", 16),
        ("B4", "Query ID Selected", 18),
        ("C4", "Reported P@5", 14),
        ("D4", "Verified Live P@5", 16),
        ("E4", "Delta Match?", 14),
        ("F4", "Facilitator Name", 20),
        ("G4", "Signature & Time", 20),
    ]
    for cell_ref, text, col_width in log_headers:
        c = ws_log[cell_ref]
        c.value = text
        c.font = font_tbl_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = cell_ref[0]
        ws_log.column_dimensions[col_letter].width = col_width

    for r_idx in range(5, 15):
        for col_idx in range(1, 8):
            c = ws_log.cell(row=r_idx, column=col_idx)
            c.border = thin_border
            if r_idx % 2 == 0:
                c.fill = fill_alt

    os.makedirs(os.path.dirname(output_xlsx_path), exist_ok=True)
    wb.save(output_xlsx_path)
    wb.close()
    print(f"[OK] Generated Excel Scorecard: {output_xlsx_path}")


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    scorecard_path = os.path.join(base_dir, "templates", "Day2_Retrieval_Scorecard_Template.xlsx")
    create_retrieval_scorecard(scorecard_path)
